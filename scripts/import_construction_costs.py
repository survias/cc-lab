from __future__ import annotations

import argparse
import re
import sqlite3
import sys
import unicodedata
from contextlib import closing
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from utils.config import (
    BACKUPS_PATH,
    CONSTRUCTION_SOURCE_PATH,
    DATABASE_PATH,
    MIGRATIONS_PATH,
)
from utils.migrations import apply_pending_migrations, create_database_backup, file_sha256


SOURCE_SHEET = "Consolidado definitivo"
EXPECTED_HEADERS = (
    "ID",
    "Fecha de Emisión de Facturación",
    "Documento N° de Factura/Vale Vista",
    "Partida",
    "Institución/Empresa",
    "Monto Neto",
    "IVA",
    "Total",
    "Monto Neto UF",
    "IVA UF",
    "Total UF",
    "Presentado en informe N°",
    "Observaciones IF",
    "Respuesta SURVIAS",
)


@dataclass(frozen=True)
class ConstructionImportResult:
    construction_import_id: int
    created: bool
    row_count: int
    report_count: int
    first_report_no: int
    last_report_no: int
    file_hash: str
    backup_path: Path | None
    migrations_applied: tuple[str, ...]


def _text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value)
    return text if text.strip() else None


def _identifier_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip() or None


def normalize_invoice_key(value: Any) -> str | None:
    original = _identifier_text(value)
    if original is None:
        return None
    normalized = unicodedata.normalize("NFKD", original).upper().strip()
    if normalized in {"N/A", "NA", "S/N", "SIN FOLIO", "NO APLICA"}:
        return None
    key = re.sub(r"[^A-Z0-9]+", "", normalized)
    return key or None


def classify_if_observation(value: Any) -> str:
    original = _text(value)
    if original is None:
        return "NO_OBSERVATION"
    equivalent = " ".join(unicodedata.normalize("NFKC", original).split()).casefold()
    return "APPROVED_EXPLICIT" if equivalent == "aprobado" else "OBSERVED"


def _number(value: Any, source_row: int, header: str) -> float:
    if value is None or (isinstance(value, str) and not value.strip()):
        return 0.0
    if isinstance(value, bool):
        raise ValueError(f"Fila {source_row}: {header} contiene un booleano.")
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("$", "").replace("UF", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError as exc:
        raise ValueError(f"Fila {source_row}: {header} no es numérico: {value!r}") from exc


def _integer(value: Any, source_row: int, header: str) -> int:
    number = _number(value, source_row, header)
    if not number.is_integer():
        raise ValueError(f"Fila {source_row}: {header} debe ser entero: {value!r}")
    return int(number)


def _date_iso(value: Any, source_row: int) -> str | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return from_excel(value).date().isoformat()
    text = str(value).strip()
    for date_format in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, date_format).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"Fila {source_row}: fecha no reconocida: {value!r}")


def read_construction_rows(source_path: Path) -> tuple[list[dict[str, Any]], list[str], int]:
    workbook = load_workbook(source_path, read_only=True, data_only=True, keep_links=False)
    try:
        if SOURCE_SHEET not in workbook.sheetnames:
            raise ValueError(f"No existe la hoja requerida: {SOURCE_SHEET}")
        sheet = workbook[SOURCE_SHEET]
        headers = tuple(sheet.cell(1, column).value for column in range(1, len(EXPECTED_HEADERS) + 1))
        if headers != EXPECTED_HEADERS:
            raise ValueError(
                "Encabezados inesperados en Consolidado definitivo. "
                f"Esperados: {EXPECTED_HEADERS}. Recibidos: {headers}."
            )

        rows: list[dict[str, Any]] = []
        mismatches: list[str] = []
        total_row = 0
        for source_row, values in enumerate(
            sheet.iter_rows(min_row=2, max_col=len(EXPECTED_HEADERS), values_only=True), start=2
        ):
            external_raw = values[0]
            if external_raw is None and str(values[1] or "").strip().casefold() == "total":
                total_row = source_row
                continue
            if all(value is None for value in values):
                continue

            external_id = _identifier_text(external_raw)
            if external_id is None:
                raise ValueError(f"Fila {source_row}: falta ID.")
            match = re.fullmatch(r"(\d+)-(\d+)", external_id)
            if not match:
                raise ValueError(f"Fila {source_row}: ID inválido: {external_id!r}")
            cost_sequence_no = int(match.group(1))
            report_no = int(match.group(2))
            reported_column = _integer(values[11], source_row, EXPECTED_HEADERS[11])
            if report_no != reported_column:
                mismatches.append(
                    f"fila {source_row}: ID {external_id} vs informe {reported_column}"
                )

            rows.append(
                {
                    "external_id": external_id,
                    "cost_sequence_no": cost_sequence_no,
                    "report_no": report_no,
                    "source_row": source_row,
                    "issue_date": _date_iso(values[1], source_row),
                    "invoice_number_reported": _identifier_text(values[2]),
                    "invoice_key": normalize_invoice_key(values[2]),
                    "description": _text(values[3]),
                    "supplier_name_reported": _text(values[4]),
                    "net_amount_clp": _number(values[5], source_row, EXPECTED_HEADERS[5]),
                    "vat_amount_clp": _number(values[6], source_row, EXPECTED_HEADERS[6]),
                    "total_amount_clp": _number(values[7], source_row, EXPECTED_HEADERS[7]),
                    "net_amount_uf": _number(values[8], source_row, EXPECTED_HEADERS[8]),
                    "vat_amount_uf": _number(values[9], source_row, EXPECTED_HEADERS[9]),
                    "total_amount_uf": _number(values[10], source_row, EXPECTED_HEADERS[10]),
                    "if_observation_raw": _text(values[12]),
                    "survias_response_raw": _text(values[13]),
                    "if_observation_class": classify_if_observation(values[12]),
                    "support_type": "PENDING_CLASSIFICATION",
                    "reconciliation_status": "PENDING_REVIEW",
                }
            )
    finally:
        workbook.close()

    if not rows:
        raise ValueError("La hoja no contiene registros de costos.")
    if total_row == 0:
        raise ValueError("No se encontró la fila final Total.")
    external_ids = [row["external_id"] for row in rows]
    if len(external_ids) != len(set(external_ids)):
        raise ValueError("La hoja contiene ID duplicados.")
    return rows, mismatches, total_row


def _existing_result(
    connection: sqlite3.Connection,
    file_hash: str,
    migrations_applied: list[str],
) -> ConstructionImportResult | None:
    row = connection.execute(
        """
        SELECT construction_import_id, row_count, first_report_no, last_report_no
        FROM construction_imports
        WHERE file_hash = ?
        """,
        (file_hash,),
    ).fetchone()
    if row is None:
        return None
    report_count = connection.execute(
        """
        SELECT COUNT(DISTINCT report_no)
        FROM construction_cost_items
        WHERE construction_import_id = ?
        """,
        (row[0],),
    ).fetchone()[0]
    return ConstructionImportResult(
        construction_import_id=int(row[0]),
        created=False,
        row_count=int(row[1]),
        report_count=int(report_count),
        first_report_no=int(row[2]),
        last_report_no=int(row[3]),
        file_hash=file_hash,
        backup_path=None,
        migrations_applied=tuple(migrations_applied),
    )


def import_construction_costs(
    source_path: Path = CONSTRUCTION_SOURCE_PATH,
    database_path: Path = DATABASE_PATH,
    migrations_path: Path = MIGRATIONS_PATH,
    backup_dir: Path = BACKUPS_PATH,
) -> ConstructionImportResult:
    source_path = source_path.expanduser().resolve()
    if not source_path.is_file():
        raise FileNotFoundError(f"No existe el consolidado: {source_path}")
    if not database_path.is_file():
        raise FileNotFoundError(f"No existe SQLite: {database_path}")

    source_hash = file_sha256(source_path)
    rows, mismatches, total_row = read_construction_rows(source_path)
    migrations_applied = apply_pending_migrations(database_path, migrations_path, backup_dir)

    with closing(sqlite3.connect(database_path)) as connection:
        connection.execute("PRAGMA foreign_keys = ON")
        existing = _existing_result(connection, source_hash, migrations_applied)
        if existing is not None:
            return existing

    backup_path = create_database_backup(database_path, backup_dir, "construction_import")
    reports = sorted({int(row["report_no"]) for row in rows})
    modified_at = datetime.fromtimestamp(source_path.stat().st_mtime).astimezone().isoformat(timespec="seconds")
    imported_at = datetime.now(timezone.utc).isoformat(timespec="seconds")
    mismatch_note = "; ".join(mismatches) if mismatches else "ninguna"
    notes = f"Fila Total excluida: {total_row}. Diferencias ID/informe: {mismatch_note}."

    insert_columns = (
        "external_id",
        "cost_sequence_no",
        "report_no",
        "source_row",
        "issue_date",
        "invoice_number_reported",
        "invoice_key",
        "description",
        "supplier_name_reported",
        "net_amount_clp",
        "vat_amount_clp",
        "total_amount_clp",
        "net_amount_uf",
        "vat_amount_uf",
        "total_amount_uf",
        "if_observation_raw",
        "survias_response_raw",
        "if_observation_class",
        "support_type",
        "reconciliation_status",
    )
    placeholders = ", ".join("?" for _ in insert_columns)
    with closing(sqlite3.connect(database_path)) as connection:
        connection.execute("PRAGMA foreign_keys = ON")
        try:
            connection.execute("BEGIN IMMEDIATE")
            connection.execute("UPDATE construction_imports SET is_active = 0 WHERE is_active = 1")
            cursor = connection.execute(
                """
                INSERT INTO construction_imports(
                    source_file_path, source_file_name, source_sheet, file_hash,
                    file_modified_at, imported_at, row_count, first_report_no,
                    last_report_no, is_active, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
                """,
                (
                    str(source_path),
                    source_path.name,
                    SOURCE_SHEET,
                    source_hash,
                    modified_at,
                    imported_at,
                    len(rows),
                    reports[0],
                    reports[-1],
                    notes,
                ),
            )
            import_id = int(cursor.lastrowid)
            connection.executemany(
                f"""
                INSERT INTO construction_cost_items(
                    construction_import_id, {', '.join(insert_columns)}
                ) VALUES (?, {placeholders})
                """,
                [
                    (import_id, *(row[column] for column in insert_columns))
                    for row in rows
                ],
            )
            connection.commit()
        except Exception:
            connection.rollback()
            raise

    return ConstructionImportResult(
        construction_import_id=import_id,
        created=True,
        row_count=len(rows),
        report_count=len(reports),
        first_report_no=reports[0],
        last_report_no=reports[-1],
        file_hash=source_hash,
        backup_path=backup_path,
        migrations_applied=tuple(migrations_applied),
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Importa de forma versionada los costos de construcción informados al MOP."
    )
    parser.add_argument("source", nargs="?", type=Path, default=CONSTRUCTION_SOURCE_PATH)
    parser.add_argument("--database", type=Path, default=DATABASE_PATH)
    args = parser.parse_args()
    result = import_construction_costs(args.source, args.database)
    action = "Importación creada" if result.created else "Archivo ya importado"
    print(f"{action}: {result.construction_import_id}")
    print(f"SHA-256: {result.file_hash}")
    print(f"Registros: {result.row_count}")
    print(
        f"Informes: {result.report_count} "
        f"({result.first_report_no}-{result.last_report_no})"
    )
    if result.migrations_applied:
        print(f"Migraciones: {', '.join(result.migrations_applied)}")
    if result.backup_path:
        print(f"Respaldo: {result.backup_path}")


if __name__ == "__main__":
    main()
