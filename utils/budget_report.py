from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import Path
import re
import unicodedata

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from utils.config import BUDGET_TEMPLATE_PATH
from utils.legacy_data import load_cost_control


ACTUAL_SHEET_PREFIX = "Cost Until "
SQLITE_SHEET = "_SQLite_Data"
SUMMARY_ROWS = {
    100: 2,
    103: 31,
    700: 35,
    500: 43,
    200: 47,
    300: 62,
    400: 129,
    800: 597,
}
CATEGORY_NAMES = {
    100: "EPC",
    103: "ITS",
    200: "Maintenance",
    300: "Operations",
    400: "SPV",
    500: "MOP",
    600: "Financing",
    700: "Insurance & Guarantee",
    800: "Tax",
}
BUDGET_CENTER_OVERRIDES = {
    2: (100, 107),
    3: (1000, 1003),
    4: (100, 107),
    5: (100, 107),
    6: (100, 107),
    7: (100, 107),
    8: (100, 107),
    9: (1000, 1001),
    10: (1000, 1001),
    11: (1000, 1001),
    12: (1000, 1001),
    13: (100, 108),
    14: (100, 108),
    15: (100, 108),
    16: (100, 108),
    17: (100, 108),
    18: (100, 108),
    22: (100, "105-106-107"),
    26: (200, 202),
    33: (200, 201),
    43: (900, 904),
}


@dataclass(frozen=True)
class BudgetExport:
    content: bytes
    filename: str
    period: pd.Timestamp
    actual_total_uf: float
    budget_total_uf: float
    matched_amount_pct: float
    unmatched_suppliers: int


def _as_int(value: object) -> int | None:
    if value is None or pd.isna(value):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _month(value: object) -> pd.Timestamp | None:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return pd.Timestamp(parsed).to_period("M").to_timestamp()


def _report_category(category: object, subcategory: object) -> int | None:
    category_code = _as_int(category)
    subcategory_code = _as_int(subcategory)
    if category_code == 100 and subcategory_code == 103:
        return 103
    if category_code == 1000:
        return 100
    if category_code == 900:
        return 400
    return category_code


def _section_row(category: object, subcategory: object) -> int:
    report_category = _report_category(category, subcategory)
    return SUMMARY_ROWS.get(report_category, 0)


def _template_subcategory(category: object, subcategory: object) -> int | None:
    category_code = _as_int(category)
    subcategory_code = _as_int(subcategory)
    if category_code == 1000:
        return {1001: 105, 1002: 108, 1003: 110}.get(subcategory_code)
    if category_code == 100:
        return {
            101: 101,
            102: 102,
            103: 103,
            104: 104,
            105: 106,
            106: 107,
            107: 108,
            108: 109,
            110: 110,
        }.get(subcategory_code)
    return subcategory_code


def _text_key(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    text = text.upper().replace("→", " ")
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    legal_terms = {
        "SOCIEDAD", "ANONIMA", "S", "A", "SPA", "LIMITADA", "LTDA", "EIRL",
        "EMPRESA", "INDIVIDUAL", "RESPONSABILIDAD", "AGENCIA", "EN", "CHILE",
    }
    words = [word for word in text.split() if word not in legal_terms]
    return " ".join(words)


def _subcategory_from_label(value: object) -> int | None:
    matches = re.findall(r"\((\d{3,4})\)", "" if value is None else str(value))
    return int(matches[-1]) if matches else None


def _actual_sheet_name(workbook) -> str:
    names = [name for name in workbook.sheetnames if name.startswith(ACTUAL_SHEET_PREFIX)]
    if not names:
        raise ValueError("La plantilla no contiene la hoja Cost Until.")
    return names[0]


def _budget_center(row: int, category: object, subcategory: object) -> tuple[object, object]:
    return BUDGET_CENTER_OVERRIDES.get(row, (category, subcategory))


def load_budget_monthly(
    template_path: Path = BUDGET_TEMPLATE_PATH,
) -> pd.DataFrame:
    workbook = load_workbook(template_path, data_only=True, read_only=False)
    sheet = workbook["Budget"]
    records: list[dict[str, object]] = []
    for column in range(6, sheet.max_column + 1):
        period = _month(sheet.cell(1, column).value)
        if period is None:
            continue
        for row in range(2, 54):
            category = _as_int(sheet.cell(row, 4).value)
            if category is None:
                continue
            subcategory = _subcategory_from_label(sheet.cell(row, 5).value)
            if subcategory is None:
                subcategory = sheet.cell(row, 5).value
            category, subcategory = _budget_center(row, category, subcategory)
            value = pd.to_numeric(sheet.cell(row, column).value, errors="coerce")
            records.append(
                {
                    "MONTH": period,
                    "CATEGORY": category,
                    "SUBCATEGORY": subcategory,
                    "REPORT_CATEGORY": _report_category(category, subcategory),
                    "ITEM": sheet.cell(row, 3).value,
                    "BUDGET_UF": 0.0 if pd.isna(value) else float(value),
                }
            )
    workbook.close()
    return pd.DataFrame.from_records(records)


def _normalize_budget_sheet(workbook) -> None:
    sheet = workbook["Budget"]
    for row, (category, subcategory) in BUDGET_CENTER_OVERRIDES.items():
        sheet.cell(row, 4).value = category
        sheet.cell(row, 5).value = subcategory


def build_actual_records(
    ledger: pd.DataFrame | None = None,
    through_month: object | None = None,
) -> pd.DataFrame:
    source = load_cost_control() if ledger is None else ledger.copy()
    actual = source[
        source["INCLUDED_IN_COST"].fillna(False)
        & source["PAID"].fillna(False)
        & source["NET-UF-F"].notna()
    ].copy()
    actual["REPORT_DATE"] = pd.to_datetime(
        actual["LAST_PAYMENT_DATE"], errors="coerce"
    ).combine_first(pd.to_datetime(actual["DATE-F"], errors="coerce"))
    actual = actual[actual["REPORT_DATE"].notna()].copy()
    actual["REPORT_MONTH"] = actual["REPORT_DATE"].dt.to_period("M").dt.to_timestamp()
    limit = _month(through_month)
    if limit is not None:
        actual = actual[actual["REPORT_MONTH"] <= limit].copy()
    actual["REPORT_CATEGORY"] = actual.apply(
        lambda row: _report_category(row["CATEGORY-F"], row["SUBCATEGORY-F"]), axis=1
    )
    actual["SECTION_ROW"] = actual.apply(
        lambda row: _section_row(row["CATEGORY-F"], row["SUBCATEGORY-F"]), axis=1
    )
    actual["NET_UF"] = pd.to_numeric(actual["NET-UF-F"], errors="coerce").fillna(0.0)
    return actual


def available_actual_months(ledger: pd.DataFrame | None = None) -> list[pd.Timestamp]:
    actual = build_actual_records(ledger)
    if actual.empty:
        return []
    start = actual["REPORT_MONTH"].min()
    end = actual["REPORT_MONTH"].max()
    return list(pd.date_range(start, end, freq="MS"))


def budget_comparison(
    ledger: pd.DataFrame | None = None,
    through_month: object | None = None,
    template_path: Path = BUDGET_TEMPLATE_PATH,
) -> pd.DataFrame:
    limit = _month(through_month)
    budget = load_budget_monthly(template_path)
    actual = build_actual_records(ledger, limit)
    if limit is not None:
        budget = budget[budget["MONTH"] <= limit]
    budget_monthly = (
        budget.groupby(["MONTH", "REPORT_CATEGORY"], dropna=False)["BUDGET_UF"]
        .sum()
        .reset_index()
    )
    actual_monthly = (
        actual.groupby(["REPORT_MONTH", "REPORT_CATEGORY"], dropna=False)["NET_UF"]
        .sum()
        .reset_index()
        .rename(columns={"REPORT_MONTH": "MONTH", "NET_UF": "ACTUAL_UF"})
    )
    comparison = budget_monthly.merge(
        actual_monthly, on=["MONTH", "REPORT_CATEGORY"], how="outer"
    ).fillna({"BUDGET_UF": 0.0, "ACTUAL_UF": 0.0})
    comparison = comparison[comparison["REPORT_CATEGORY"].notna()].copy()
    comparison["REPORT_CATEGORY"] = comparison["REPORT_CATEGORY"].astype(int)
    comparison["CATEGORY_NAME"] = comparison["REPORT_CATEGORY"].map(CATEGORY_NAMES)
    comparison["BALANCE_UF"] = comparison["BUDGET_UF"] - comparison["ACTUAL_UF"]
    return comparison.sort_values(["MONTH", "REPORT_CATEGORY"])


def _template_rows(sheet) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    section = 0
    for row in range(2, 598):
        if row in SUMMARY_ROWS.values():
            section = row
            continue
        supplier = sheet.cell(row, 5).value
        if not isinstance(supplier, str) or "→" not in supplier:
            continue
        rows.append(
            {
                "row": row,
                "section": section,
                "subcategory": _subcategory_from_label(sheet.cell(row, 4).value),
                "supplier": supplier.split("→", 1)[1].strip(),
                "supplier_key": _text_key(supplier),
            }
        )
    return rows


def _match_template_row(
    row: pd.Series,
    candidates: list[dict[str, object]],
) -> tuple[int, str]:
    section = int(row["SECTION_ROW"])
    if section == 0:
        return 0, "NO_SECTION"
    category = _as_int(row["CATEGORY-F"])
    subcategory = _as_int(row["SUBCATEGORY-F"])
    if category == 100 and subcategory == 109:
        return 18, "CENTER_RULE"
    target_subcategory = _template_subcategory(category, subcategory)
    section_rows = [item for item in candidates if item["section"] == section]
    same_center = [
        item for item in section_rows
        if target_subcategory is not None and item["subcategory"] == target_subcategory
    ]
    pool = same_center or section_rows
    if not pool:
        return 0, "NO_TEMPLATE_ROW"
    if len(pool) == 1:
        return int(pool[0]["row"]), "CENTER_RULE"
    supplier_key = _text_key(row["SUPPLIER-F"])
    exact = [item for item in pool if item["supplier_key"] == supplier_key]
    if len(exact) == 1:
        return int(exact[0]["row"]), "SUPPLIER_EXACT"
    scored = [
        (
            SequenceMatcher(None, supplier_key, str(item["supplier_key"])).ratio(),
            int(item["row"]),
        )
        for item in pool
    ]
    score, matched_row = max(scored, default=(0.0, 0))
    if score >= 0.82:
        return matched_row, "SUPPLIER_SIMILAR"
    return 0, "UNMATCHED_SUPPLIER"


def _extend_month_columns(sheet, through_month: pd.Timestamp) -> list[tuple[int, pd.Timestamp]]:
    columns: list[tuple[int, pd.Timestamp]] = []
    for column in range(7, sheet.max_column + 1):
        period = _month(sheet.cell(1, column).value)
        if period is None:
            continue
        columns.append((column, period))
    if not columns:
        raise ValueError("La hoja Cost Until no contiene columnas mensuales.")
    last_column, last_period = columns[-1]
    while last_period < through_month:
        new_column = last_column + 1
        if new_column > sheet.max_column:
            raise ValueError("La plantilla no tiene columnas libres para agregar nuevos meses.")
        last_period = last_period + pd.offsets.MonthBegin(1)
        source_letter = get_column_letter(last_column)
        target_letter = get_column_letter(new_column)
        sheet.column_dimensions[target_letter].width = sheet.column_dimensions[source_letter].width
        sheet.column_dimensions[target_letter].hidden = False
        for row in range(1, 598):
            source = sheet.cell(row, last_column)
            target = sheet.cell(row, new_column)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)
        sheet.cell(1, new_column).value = last_period.strftime("%Y-%m")
        columns.append((new_column, last_period))
        last_column = new_column
    return columns


def _replace_sheet_references(workbook, old_name: str, new_name: str) -> None:
    old_quoted = f"'{old_name}'!"
    new_quoted = f"'{new_name}'!"
    old_plain = f"{old_name}!"
    new_plain = f"{new_name}!"
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = cell.value.replace(old_quoted, new_quoted).replace(
                        old_plain, new_plain
                    )
    for defined_name in workbook.defined_names.values():
        if isinstance(defined_name.attr_text, str):
            defined_name.attr_text = defined_name.attr_text.replace(
                old_quoted, new_quoted
            ).replace(old_plain, new_plain)


def _write_sqlite_sheet(workbook, actual: pd.DataFrame) -> None:
    if SQLITE_SHEET in workbook.sheetnames:
        del workbook[SQLITE_SHEET]
    sheet = workbook.create_sheet(SQLITE_SHEET)
    headers = [
        "REPORT_MONTH", "REPORT_DATE", "RUT", "SUPPLIER", "DOCUMENT_TYPE", "FOLIO",
        "CATEGORY", "SUBCATEGORY", "NET_UF", "SOURCE", "DOCUMENT_ID", "PAYMENT_ID",
        "SECTION_ROW", "TEMPLATE_ROW", "MATCH_STATUS",
    ]
    sheet.append(headers)
    for _, item in actual.iterrows():
        sheet.append(
            [
                item["REPORT_MONTH"].to_pydatetime(),
                item["REPORT_DATE"].to_pydatetime(),
                item["RUT_COMPLETO"],
                item["SUPPLIER-F"],
                _as_int(item["DOCUMENT TYPE"]),
                item["INVOICE-F"],
                _as_int(item["CATEGORY-F"]),
                _as_int(item["SUBCATEGORY-F"]),
                float(item["NET_UF"]),
                item["SOURCE_KIND"],
                _as_int(item["DOCUMENT_ID"]),
                _as_int(item["PAYMENT_ID"]),
                int(item["SECTION_ROW"]),
                int(item["TEMPLATE_ROW"]),
                item["MATCH_STATUS"],
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:O{max(sheet.max_row, 2)}"
    sheet.column_dimensions["A"].width = 13
    sheet.column_dimensions["B"].width = 13
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 48
    sheet.column_dimensions["F"].width = 16
    sheet.column_dimensions["I"].width = 14
    for cell in sheet["A"][1:]:
        cell.number_format = "yyyy-mm"
    for cell in sheet["B"][1:]:
        cell.number_format = "dd-mm-yyyy"
    for cell in sheet["I"][1:]:
        cell.number_format = '#,##0.00 "UF"'
    sheet.sheet_state = "hidden"


def build_budget_workbook(
    through_month: object,
    ledger: pd.DataFrame | None = None,
    template_path: Path = BUDGET_TEMPLATE_PATH,
) -> BudgetExport:
    period = _month(through_month)
    if period is None:
        raise ValueError("Debe indicar un mes valido para generar el reporte.")
    actual = build_actual_records(ledger, period)
    workbook = load_workbook(template_path, data_only=False, keep_links=True)
    _normalize_budget_sheet(workbook)
    old_name = _actual_sheet_name(workbook)
    actual_sheet = workbook[old_name]
    template_rows = _template_rows(actual_sheet)
    matches = actual.apply(
        lambda row: _match_template_row(row, template_rows),
        axis=1,
        result_type="expand",
    )
    actual["TEMPLATE_ROW"] = matches[0].astype(int)
    actual["MATCH_STATUS"] = matches[1]
    _write_sqlite_sheet(workbook, actual)

    month_columns = _extend_month_columns(actual_sheet, period)
    detail_totals = (
        actual[actual["TEMPLATE_ROW"] > 0]
        .groupby(["TEMPLATE_ROW", "REPORT_MONTH"])["NET_UF"]
        .sum()
        .to_dict()
    )
    detail_rows = {int(item["row"]) for item in template_rows}
    last_column = month_columns[-1][0]
    last_letter = get_column_letter(last_column)
    for column, month_value in month_columns:
        date_formula = f"DATE({month_value.year},{month_value.month},1)"
        for summary_row in SUMMARY_ROWS.values():
            actual_sheet.cell(summary_row, column).value = (
                f"=SUMIFS('{SQLITE_SHEET}'!$I:$I,'{SQLITE_SHEET}'!$M:$M,"
                f"{summary_row},'{SQLITE_SHEET}'!$A:$A,{date_formula})"
            )
        for detail_row in detail_rows:
            actual_sheet.cell(detail_row, column).value = float(
                detail_totals.get((detail_row, month_value), 0.0)
            )
    for row in set(SUMMARY_ROWS.values()) | detail_rows:
        actual_sheet.cell(row, 6).value = f"=SUM(G{row}:{last_letter}{row})"
    actual_sheet.auto_filter.ref = f"C1:{last_letter}597"

    new_name = f"{ACTUAL_SHEET_PREFIX}{period:%m-%y}"
    if new_name != old_name:
        _replace_sheet_references(workbook, old_name, new_name)
        actual_sheet.title = new_name
    if "Summary" in workbook.sheetnames:
        workbook["Summary"]["E3"] = f"UNTIL {period.strftime('%B %Y').upper()}"
    if "Balance" in workbook.sheetnames:
        workbook["Balance"]["D1"] = period.strftime("%B %Y").upper()
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    budget = load_budget_monthly(template_path)
    budget_total = float(budget.loc[budget["MONTH"] <= period, "BUDGET_UF"].sum())
    actual_total = float(actual["NET_UF"].sum())
    absolute_total = float(actual["NET_UF"].abs().sum())
    matched_amount = float(
        actual.loc[actual["TEMPLATE_ROW"] > 0, "NET_UF"].abs().sum()
    )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return BudgetExport(
        content=output.getvalue(),
        filename=f"Cost vs Budget Analysis_{period:%m%y}.xlsx",
        period=period,
        actual_total_uf=actual_total,
        budget_total_uf=budget_total,
        matched_amount_pct=matched_amount / absolute_total if absolute_total else 1.0,
        unmatched_suppliers=int(
            actual.loc[actual["TEMPLATE_ROW"] == 0, "SUPPLIER-F"].nunique()
        ),
    )
